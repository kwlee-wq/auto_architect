"""
AutoArchitect - 샘플 데이터 생성 스크립트
실행: python scripts/create_samples.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path


def apply_header_style(cell):
    """헤더 스타일"""
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_data_style(cell):
    """데이터 스타일"""
    cell.alignment = Alignment(horizontal="left", vertical="center")


def create_basic_template():
    """기본형 엑셀 템플릿 생성"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # CONFIG 시트
    ws = wb.create_sheet("CONFIG")
    headers = ["항목", "값", "설명"]
    for col, header in enumerate(headers, 1):
        apply_header_style(ws.cell(1, col, header))

    data = [
        ["다이어그램명", "시스템 구성도", "구성도 제목"],
        ["캔버스너비", 1200, "픽셀 단위"],
        ["캔버스높이", 800, "픽셀 단위"],
        ["레이아웃패턴", "수평레이어스택", "수평레이어스택/좌우분할/중앙허브형/좌우파이프라인"],
        ["여백비율", 15, "컴포넌트 간 여백 (10-30)"]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            apply_data_style(ws.cell(row_idx, col_idx, value))

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50

    # LAYERS 시트
    ws = wb.create_sheet("LAYERS")
    headers = ["레이어ID", "레이어명", "높이%", "배경색", "테두리색"]
    for col, header in enumerate(headers, 1):
        apply_header_style(ws.cell(1, col, header))

    data = [
        ["L1", "Application Layer", 25, "하늘색", "진한파랑"],
        ["L2", "Service Layer", 25, "연두색", "진한녹색"],
        ["L3", "Data Layer", 50, "주황색", "진한주황"]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            apply_data_style(ws.cell(row_idx, col_idx, value))

    # COMPONENTS 시트
    ws = wb.create_sheet("COMPONENTS")
    headers = ["ID", "컴포넌트명", "레이어ID", "타입", "너비", "아이콘", "텍스트크기"]
    for col, header in enumerate(headers, 1):
        apply_header_style(ws.cell(1, col, header))

    data = [
        ["C1", "웹 서버", "L1", "서비스", 2, "web", "중간"],
        ["C2", "API 서버", "L2", "서비스", 3, "api", "중간"],
        ["C3", "데이터베이스", "L3", "데이터베이스", 2, "database", "중간"],
        ["C4", "캐시 서버", "L3", "서비스", 2, "storage", "중간"]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            apply_data_style(ws.cell(row_idx, col_idx, value))

    # CONNECTIONS 시트
    ws = wb.create_sheet("CONNECTIONS")
    headers = ["출발ID", "도착ID", "연결타입", "라벨", "선스타일"]
    for col, header in enumerate(headers, 1):
        apply_header_style(ws.cell(1, col, header))

    data = [
        ["C1", "C2", "데이터흐름", "REST API", "실선"],
        ["C2", "C3", "데이터흐름", "SQL", "실선"],
        ["C2", "C4", "데이터흐름", "Cache", "점선"]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            apply_data_style(ws.cell(row_idx, col_idx, value))

    # GUIDE 시트
    ws = wb.create_sheet("GUIDE")
    ws.cell(1, 1, "📖 AutoArchitect 사용 가이드")
    ws.cell(3, 1, "1. CONFIG: 기본 설정")
    ws.cell(4, 1, "2. LAYERS: 레이어 정의 (높이% 합계 = 100)")
    ws.cell(5, 1, "3. COMPONENTS: 컴포넌트 정의")
    ws.cell(6, 1, "4. CONNECTIONS: 연결 관계")

    output_path = Path("templates/excel_template.xlsx")
    wb.save(output_path)
    print(f"✅ 기본 템플릿 생성: {output_path}")


def create_nested_sample():
    """계층형 샘플 데이터 생성 (우체국 스타일)"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # CONFIG
    ws = wb.create_sheet("CONFIG")
    headers = ["항목", "값"]
    for col, header in enumerate(headers, 1):
        apply_header_style(ws.cell(1, col, header))

    data = [
        ["다이어그램명", "빅데이터 플랫폼"],
        ["캔버스너비", 1400],
        ["캔버스높이", 900]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            apply_data_style(ws.cell(row_idx, col_idx, value))

    # LAYERS
    ws = wb.create_sheet("LAYERS")
    headers = ["레이어ID", "레이어명", "순서", "배경색", "높이%"]
    for col, header in enumerate(headers, 1):
        apply_header_style(ws.cell(1, col, header))

    data = [
        ["L1", "Service Layer", 1, "연회색", 15],
        ["L2", "Application Layer", 2, "흰색", 85]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            apply_data_style(ws.cell(row_idx, col_idx, value))

    # BOXES
    ws = wb.create_sheet("BOXES")
    headers = ["박스ID", "박스명", "부모ID", "행번호", "Y%", "높이%", "배경색", "테두리색", "폰트크기"]
    for col, header in enumerate(headers, 1):
        apply_header_style(ws.cell(1, col, header))

    data = [
        ["B1", "공통기능", "L1", 1, 20, 65, "흰색", "회색", 10],
        ["B2", "모니터링", "L1", 1, 20, 65, "흰색", "회색", 10],
        ["B3", "시각화", "L1", 1, 20, 65, "흰색", "회색", 10],
        ["B4", "Interface", "L2", 1, 8, 88, "연회색", "회색", 11],
        ["B5", "Data Lake", "L2", 1, 8, 88, "연회색", "회색", 11],
        ["B6", "분석 플랫폼", "L2", 1, 8, 88, "연회색", "회색", 11]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            apply_data_style(ws.cell(row_idx, col_idx, value))

    # COMPONENTS
    ws = wb.create_sheet("COMPONENTS")
    headers = ["ID", "컴포넌트명", "부모ID", "행번호", "Y%", "높이%", "폰트크기", "타입"]
    for col, header in enumerate(headers, 1):
        apply_header_style(ws.cell(1, col, header))

    data = [
        ["C1", "JDBC Interface", "B4", 1, 10, 25, 9, "단일박스"],
        ["C2", "Batch Interface", "B4", 1, 40, 25, 9, "단일박스"],
        ["C3", "Hadoop Cluster", "B5", 1, 10, 80, 10, "클러스터"],
        ["C4", "ML Modeler", "B6", 1, 10, 40, 9, "서비스"],
        ["C5", "Auto ML", "B6", 1, 55, 40, 9, "서비스"]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            apply_data_style(ws.cell(row_idx, col_idx, value))

    # CONNECTIONS
    ws = wb.create_sheet("CONNECTIONS")
    headers = ["출발ID", "도착ID", "연결타입", "라벨", "선스타일"]
    for col, header in enumerate(headers, 1):
        apply_header_style(ws.cell(1, col, header))

    data = [
        ["B4", "B5", "데이터흐름", "", "실선"],
        ["B5", "B6", "데이터흐름", "", "실선"]
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            apply_data_style(ws.cell(row_idx, col_idx, value))

    output_path = Path("templates/nested_sample.xlsx")
    wb.save(output_path)
    print(f"✅ 계층형 샘플 생성: {output_path}")


def main():
    """모든 템플릿 및 샘플 생성"""
    print("🔧 AutoArchitect 템플릿 생성 시작...")

    # templates 디렉토리 생성
    Path("templates").mkdir(exist_ok=True)

    create_basic_template()
    create_nested_sample()

    print("✅ 모든 템플릿 생성 완료!")


if __name__ == "__main__":
    main()
