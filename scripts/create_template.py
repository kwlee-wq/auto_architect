"""
엑셀 템플릿 파일 생성 스크립트
실행: python scripts/create_template.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path


def create_excel_template():
    """엑셀 템플릿 파일 생성"""

    # 워크북 생성
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # 각 시트 생성
    create_config_sheet(wb)
    create_layers_sheet(wb)
    create_components_sheet(wb)
    create_sub_components_sheet(wb)
    create_connections_sheet(wb)
    create_groups_sheet(wb)
    create_guide_sheet(wb)

    # 파일 저장
    output_path = Path("../templates/excel_template.xlsx")
    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)

    print(f"✅ 템플릿 생성 완료: {output_path}")


def apply_header_style(cell):
    """헤더 셀 스타일 적용"""
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def apply_data_style(cell, bg_color="FFFFFF"):
    """데이터 셀 스타일 적용"""
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )


def create_config_sheet(wb):
    """CONFIG 시트 생성"""
    ws = wb.create_sheet("CONFIG")

    # 헤더
    headers = ["항목", "값", "설명"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        apply_header_style(cell)

    # 데이터
    config_data = [
        ["다이어그램명", "우체국 빅데이터 플랫폼", "구성도 제목"],
        ["캔버스너비", 1200, "픽셀 단위 (권장: 1200-1600)"],
        ["캔버스높이", 800, "픽셀 단위 (권장: 800-1000)"],
        ["레이아웃패턴", "수평레이어스택", "수평레이어스택/좌우분할/중앙허브형/좌우파이프라인"],
        ["여백비율", 15, "컴포넌트 간 여백 (10-30)"]
    ]

    for row_idx, row_data in enumerate(config_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            apply_data_style(cell)

    # 레이아웃패턴 드롭다운
    dv_layout = DataValidation(
        type="list",
        formula1='"수평레이어스택,좌우분할,중앙허브형,좌우파이프라인"',
        allow_blank=False
    )
    ws.add_data_validation(dv_layout)
    dv_layout.add(ws['B5'])

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50


def create_layers_sheet(wb):
    """LAYERS 시트 생성"""
    ws = wb.create_sheet("LAYERS")

    # 헤더
    headers = ["레이어ID", "레이어명", "높이%", "배경색", "테두리색"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        apply_header_style(cell)

    # 샘플 데이터
    sample_data = [
        ["L1", "Application Layer", 20, "하늘색", "진한파랑"],
        ["L2", "Service Layer", 25, "연두색", "진한녹색"],
        ["L3", "Data Lake", 30, "주황색", "진한주황"],
        ["L4", "Infrastructure", 25, "회색", "진한회색"]
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)

            # 배경색 미리보기
            if col_idx == 4:  # 배경색 컬럼
                color_map = {
                    "하늘색": "E3F2FD",
                    "연두색": "E8F5E9",
                    "주황색": "FFE0B2",
                    "회색": "F5F5F5"
                }
                bg_color = color_map.get(value, "FFFFFF")
                apply_data_style(cell, bg_color)
            else:
                apply_data_style(cell)

    # 드롭다운 설정
    color_options = "하늘색,연두색,주황색,회색,흰색,노란색,분홍색,보라색,파란색,녹색"
    border_options = "진한파랑,진한녹색,진한주황,진한회색,진한빨강,진한보라,검정"

    dv_bg_color = DataValidation(type="list", formula1=f'"{color_options}"', allow_blank=False)
    dv_border_color = DataValidation(type="list", formula1=f'"{border_options}"', allow_blank=False)

    ws.add_data_validation(dv_bg_color)
    ws.add_data_validation(dv_border_color)

    # 드롭다운 적용 범위 (2행부터 100행까지)
    dv_bg_color.add(f'D2:D100')
    dv_border_color.add(f'E2:E100')

    # 높이% 숫자 검증 (0-100)
    dv_height = DataValidation(
        type="whole",
        operator="between",
        formula1=0,
        formula2=100,
        allow_blank=False,
        errorTitle="잘못된 값",
        error="0에서 100 사이의 숫자를 입력하세요"
    )
    ws.add_data_validation(dv_height)
    dv_height.add('C2:C100')

    # 컬럼 너비
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15


def create_components_sheet(wb):
    """COMPONENTS 시트 생성"""
    ws = wb.create_sheet("COMPONENTS")

    # 헤더
    headers = ["ID", "컴포넌트명", "레이어ID", "타입", "너비", "아이콘", "텍스트크기"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        apply_header_style(cell)

    # 샘플 데이터
    sample_data = [
        ["C1", "빅데이터포털", "L2", "단일박스", 3, "portal", "중간"],
        ["C2", "Hadoop Cluster", "L3", "클러스터", 4, "hadoop", "중간"],
        ["C3", "Kafka", "L4", "서비스", 2, "kafka", "작음"],
        ["C4", "Spark", "L4", "서비스", 2, "spark", "작음"],
        ["C5", "데이터베이스", "L3", "데이터베이스", 2, "database", "중간"]
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            apply_data_style(cell)

    # 드롭다운 설정
    type_options = "단일박스,클러스터,서비스,데이터베이스,저장소,문서"
    icon_options = "portal,hadoop,kafka,spark,kubernetes,database,storage,api,web,mobile,server,cloud,network,security,monitor"
    text_size_options = "작음,중간,큼,아주큼"

    dv_type = DataValidation(type="list", formula1=f'"{type_options}"', allow_blank=False)
    dv_icon = DataValidation(type="list", formula1=f'"{icon_options}"', allow_blank=True)
    dv_text_size = DataValidation(type="list", formula1=f'"{text_size_options}"', allow_blank=False)

    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_icon)
    ws.add_data_validation(dv_text_size)

    dv_type.add('D2:D100')
    dv_icon.add('F2:F100')
    dv_text_size.add('G2:G100')

    # 너비 검증 (1-5)
    dv_width = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=5,
        allow_blank=False,
        errorTitle="잘못된 값",
        error="1에서 5 사이의 숫자를 입력하세요"
    )
    ws.add_data_validation(dv_width)
    dv_width.add('E2:E100')

    # 레이어ID 참조 검증 (LAYERS 시트 참조)
    # Note: 실제로는 INDIRECT 함수로 동적 참조하는게 좋지만, 여기서는 단순화

    # 컬럼 너비
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12


def create_sub_components_sheet(wb):
    """SUB_COMPONENTS 시트 생성"""
    ws = wb.create_sheet("SUB_COMPONENTS")

    # 헤더
    headers = ["부모ID", "서브컴포넌트명", "순서"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        apply_header_style(cell)

    # 샘플 데이터 (Hadoop Cluster 내부 컴포넌트)
    sample_data = [
        ["C2", "HDFS", 1],
        ["C2", "YARN", 2],
        ["C2", "Hive", 3],
        ["C2", "Spark", 4]
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            apply_data_style(cell)

    # 순서 검증 (1-20)
    dv_order = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=20,
        allow_blank=False
    )
    ws.add_data_validation(dv_order)
    dv_order.add('C2:C100')

    # 컬럼 너비
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10


def create_connections_sheet(wb):
    """CONNECTIONS 시트 생성"""
    ws = wb.create_sheet("CONNECTIONS")

    # 헤더
    headers = ["출발ID", "도착ID", "연결타입", "라벨", "선스타일"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        apply_header_style(cell)

    # 샘플 데이터
    sample_data = [
        ["C1", "C2", "데이터흐름", "REST API", "실선"],
        ["C3", "C4", "스트림", "실시간", "굵은실선"],
        ["C2", "C5", "배치", "Daily", "점선"],
        ["C4", "C5", "데이터흐름", "저장", "실선"]
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            apply_data_style(cell)

    # 드롭다운 설정
    conn_type_options = "데이터흐름,양방향,스트림,배치"
    line_style_options = "실선,점선,굵은실선,이중선"

    dv_conn_type = DataValidation(type="list", formula1=f'"{conn_type_options}"', allow_blank=False)
    dv_line_style = DataValidation(type="list", formula1=f'"{line_style_options}"', allow_blank=False)

    ws.add_data_validation(dv_conn_type)
    ws.add_data_validation(dv_line_style)

    dv_conn_type.add('C2:C100')
    dv_line_style.add('E2:E100')

    # 컬럼 너비
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15


def create_groups_sheet(wb):
    """GROUPS 시트 생성"""
    ws = wb.create_sheet("GROUPS")

    # 헤더
    headers = ["그룹ID", "그룹명", "포함컴포넌트(IDs)", "테두리스타일", "배경투명도"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        apply_header_style(cell)

    # 샘플 데이터
    sample_data = [
        ["G1", "보안영역", "C1,C5", "빨간점선", "10%"],
        ["G2", "분석영역", "C3,C4", "파란실선", "5%"]
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            apply_data_style(cell)

    # 드롭다운 설정
    border_style_options = "빨간점선,파란실선,녹색점선,검정실선"
    transparency_options = "5%,10%,15%,20%"

    dv_border = DataValidation(type="list", formula1=f'"{border_style_options}"', allow_blank=False)
    dv_transparency = DataValidation(type="list", formula1=f'"{transparency_options}"', allow_blank=False)

    ws.add_data_validation(dv_border)
    ws.add_data_validation(dv_transparency)

    dv_border.add('D2:D100')
    dv_transparency.add('E2:E100')

    # 컬럼 너비
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15


def create_guide_sheet(wb):
    """GUIDE 시트 생성"""
    ws = wb.create_sheet("GUIDE")

    # 제목
    title_cell = ws.cell(1, 1, "📖 AutoArchitect 엑셀 템플릿 작성 가이드")
    title_cell.font = Font(bold=True, size=16, color="1F4E78")
    ws.merge_cells('A1:E1')

    # 가이드 내용
    guide_content = [
        ["", "", "", "", ""],
        ["📋 시트 구성", "", "", "", ""],
        ["시트명", "설명", "필수여부", "예시", ""],
        ["CONFIG", "다이어그램 기본 설정", "필수", "다이어그램명, 캔버스 크기 등", ""],
        ["LAYERS", "레이어(계층) 정의", "필수", "Application, Service, Data 등", ""],
        ["COMPONENTS", "컴포넌트(구성요소) 정의", "필수", "서버, DB, 서비스 등", ""],
        ["SUB_COMPONENTS", "클러스터 내부 컴포넌트", "선택", "Hadoop 내부의 HDFS, YARN 등", ""],
        ["CONNECTIONS", "컴포넌트 간 연결", "선택", "API 호출, 데이터 흐름 등", ""],
        ["GROUPS", "컴포넌트 그룹화", "선택", "보안영역, 분석영역 등", ""],
        ["", "", "", "", ""],
        ["✏️ 작성 순서", "", "", "", ""],
        ["1. CONFIG 시트", "다이어그램 이름과 크기 설정", "", "", ""],
        ["2. LAYERS 시트", "레이어를 위에서 아래 순서로 정의 (높이% 합계 100%)", "", "", ""],
        ["3. COMPONENTS 시트", "각 레이어에 속할 컴포넌트 정의", "", "", ""],
        ["4. SUB_COMPONENTS", "(선택) 클러스터 타입 컴포넌트의 내부 요소", "", "", ""],
        ["5. CONNECTIONS", "컴포넌트 간 연결 관계 정의", "", "", ""],
        ["6. GROUPS", "(선택) 여러 컴포넌트를 시각적으로 묶기", "", "", ""],
        ["", "", "", "", ""],
        ["⚠️ 주의사항", "", "", "", ""],
        ["• ID는 중복되지 않게 작성 (L1, L2, C1, C2 등)", "", "", "", ""],
        ["• 레이어ID, 컴포넌트ID는 다른 시트에서 참조하므로 정확히 입력", "", "", "", ""],
        ["• 높이% 합계는 100%에 가깝게 (±5% 허용)", "", "", "", ""],
        ["• 드롭다운 목록에서 선택 (직접 입력하지 말 것)", "", "", "", ""],
        ["• 컴포넌트명은 50자 이내 권장", "", "", "", ""],
        ["", "", "", "", ""],
        ["💡 팁", "", "", "", ""],
        ["• 레이어는 3-5개가 적당 (너무 많으면 복잡)", "", "", "", ""],
        ["• 컴포넌트는 레이어당 5-8개 권장", "", "", "", ""],
        ["• 연결은 필요한 것만 (너무 많으면 가독성 저하)", "", "", "", ""],
        ["• 너비(1-5)는 컴포넌트 중요도에 따라 조절", "", "", "", ""],
        ["", "", "", "", ""],
        ["🎨 레이아웃 패턴", "", "", "", ""],
        ["수평레이어스택", "가장 일반적, 계층 구조 명확", "권장", "", ""],
        ["좌우분할", "소스→타겟 데이터 흐름 강조", "", "", ""],
        ["중앙허브형", "중앙 시스템 중심 구조", "", "", ""],
        ["좌우파이프라인", "프로세스 단계별 흐름", "", "", ""],
        ["", "", "", "", ""],
        ["📞 문의", "", "", "", ""],
        ["문제가 있거나 개선사항이 있으면 이슈를 등록해주세요!", "", "", "", ""],
    ]

    current_row = 3
    for row_data in guide_content:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(current_row, col_idx, value)

            # 섹션 제목 스타일
            if value and value.startswith(("📋", "✏️", "⚠️", "💡", "🎨", "📞")):
                cell.font = Font(bold=True, size=12, color="1F4E78")

            # 시트명 강조
            elif col_idx == 1 and value in ["CONFIG", "LAYERS", "COMPONENTS", "SUB_COMPONENTS", "CONNECTIONS", "GROUPS"]:
                cell.font = Font(bold=True, color="0066CC")

        current_row += 1

    # 컬럼 너비
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 10


if __name__ == "__main__":
    print("🔧 엑셀 템플릿 생성 시작...")
    create_excel_template()
    print("✅ 완료!")